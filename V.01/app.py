from flask import Flask, request, render_template, send_file
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import time




app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        input1 = request.form['input1']
        input2 = request.form['input2']
        input3 = request.form['input3']
        input4 = request.form['input4']
        input5 = request.form['input5']
        input6 = request.form['input6']
        input7 = request.form['input7']
        input8 = request.form['input8']
        input9 = request.form['input9']
        input10 = request.form['input10']
        input11 = request.form['input11']
        input12 = request.form['input12']
        
        
        
        
                                    
        # Write the data to an Excel spreadsheet
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "EEL"
        
        globallist = [["site_name","DiaryGold"],["project_number","j10090"],["engineer_name",input3],["machine_type","SRT"],["design_air_temperature_C", 4],["total_evaporator_count",int(input6)],["total_fan_count" , int(input7)],["largest_fan_kw", 2.2],["VSD_DOL_selection","DOL"],["infeed_conveyor_count",1],["outfeed_conveyor_count",2]] # Not using this list just yet.
        locallist = ["io_start","local_name","area_","equipment_ID","device_ID","device_description"]

        i = 1      
        item_number = 1
        n = 51
        
        def addrow(item_number,engineer_name,area,equipment_ID,equipment_count,device_ID,device_description):  
             localname = {
        "Item" : item_number,	
        "ChangeLog" : engineer_name,	
        "ChangeDate" : datetime.now(),
        "Area" : area ,	
        "EquipmentID" : equipment_ID,
        "Equipment" : equipment_count,
        "DeviceID" : device_ID,
        "Device" : equipment_ID,	
        "DeviceTag" : area+equipment_ID+equipment_count+device_ID,	
        "DeviceDescription" : device_description,	
        "DeviceDescription2" : "",	
        "DesignNotes" : "IO-Link, 20m Max Cable Length",
        "LocationOriginField" : "CP02",
        "RatedkW" : "",	
        "LoadType" : "",	
        "EStopZone" : "",	
        "AllocatedAddressPLC_Format" : "",	
        "InputOutput" : "IO-Link",	
        "Ethernet" : "",	
        "IOLinkDI" : "1", 
        "IOLinkDO" : "",	
        "PointIODI" : "",	
        "PointIODO" : "",
        "PointIOSafeDI" : "",	
        "PointIOSafeDO" : "",	
        "PointIOAI" : "",	
        "PointIOAO" : "",	
        "LocalDI" : "",	
        "LocalDO" : "",	
        "LocalAI" : "",
        "LocalAO" : "",	
        "LocalHSCProcurementStatus" : "",	
        "Part1_Sensor" : "",	
        "Part2_Mounting" : "",	
        "Part3_": "",
        "Part_4_CablePlug": "",
        "Part_5_CablePlug": "",
        "Comments1" : "",
        "Part1_Protection":"",
        "Part2_Switchgear": "",
        "Part3_" : "",	
        "Part4_" : "",	
        "Part5_Enclosure" : "",
        "Comments2" : "",	
        "Part1_CableMarker" : "",
        "Part2_CableType" : "",
        "Part3_CableLength" : "",
        "Conductor_Size" : "",
        "Part4_MotorIsolator" : "",	
        "Part5_Enclosure" : "",	
        "Comments3" : "",

        }
             values = list(localname.values())[:n]
             ws1.append(values)
             
              
        def addtitle(title):
            ws1.append ([item_number,title])
            titlelocation = item_number
            fontchange = 'B'+str(titlelocation)
            cell = ws1[fontchange]
            cell.font = Font(name='Calibri',
                        size=14,
                        bold=True,
                        italic=True,
                        underline='single'
                        )
            cell.fill = PatternFill(start_color="afcafa", end_color="afcafa", fill_type="solid")
  
        addtitle("Refrigeration")
        item_number += 1
        #Refrigeration Interface
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTR"+"_"],["equipment_ID","INT"],["equipment_count",""],["device_ID",""],["device_description","Tunnel Evaporator Temperatures"]]
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"_"+"DC"+ locallist[6][1],"Refrigeration Interface")
        item_number += 1
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"_"+"DI01"+locallist[6][1],"Refrigeration Interface DC Supply")
        item_number += 1
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"_"+"DI02"+locallist[6][1],"Refrigeration Interface - Faulted")
        item_number += 1
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"_"+"DO01"+locallist[6][1],"Refrigeration Interface - Spare Input")
        item_number += 1
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"_"+"DO02"+locallist[6][1],"Refrigeration Interface - Start Request")
        item_number += 1
        #Tunnel Evaporator Temperatures     
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_","CTR"+"_"],["equipment_ID","EV"],["equipment_count",""],["device_ID","TT"],["device_description","Tunnel Evaporator Temperatures"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1

        i = 1
        j = 1
        k = 1

        while i <= globallist[5][1] / 2:
            for j in range(1, 3):
                print(" Inner loop: ", j)
                if k == 1:
                    onoff = "On"
                else:
                    onoff = "Off"
                locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_","CTR"+"_"],["equipment_ID","EV"],["equipment_count",str(k)],["device_ID","TT"+str(j)],["device_description","Tunnel Evaporator "+str(k)+" Air "+onoff+" Temperature"]]
                addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1]) 
                k += 1
                item_number += 1    
            i += 1
        

        i = 1 
        #Evaporator Fans (Master Line)
        addtitle("Evaporator Fans")
        item_number += 1

        i = 1
        j = 1
        k = 1
        #Evaporator Fans (Fans Line)
            
        while i <= globallist[6][1] / 2:
            for j in range(1, 3):
                print(" Inner loop: ", j)
                locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_","CTR"+"_"],["equipment_ID","FN"],["equipment_count",str(k)],["device_ID","_"+"MT"],["device_description","Evaporator Fan"+str(k)+" Motor"]]
                addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1]) 
                k += 1
                item_number += 1
            i += 1

        i = 1
        addtitle("HYDRAULICS")
        item_number += 1 
        ##HYDRAULICS (Title)
        locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_","CTH"+"_"],["equipment_ID",""],["equipment_count",""],["device_ID",""],["device_description","HYDRAULICS"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        ##HYDRAULICS (IO Master)
        locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_","CTH"+"_"],["equipment_ID","HPU"],["equipment_count",""],["device_ID","AL"+str(i)],["device_description","Hydraulic Power Unit IO-Link Master"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        while i <= 8:
            
            print(" Inner loop: ", j)
            locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_",""+"_"],["equipment_ID",""],["equipment_count",""],["device_ID",""],["device_description","IO-Link Master X0"+str(i)]]
            addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1]) 
            i += 1
            item_number += 1
        i = 1 
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","HPU"],["equipment_count",""],["device_ID",""],["device_description","Tunnel Evaporator Temperatures"]]
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],"DIS"+ locallist[6][1],"Hydraulic Power Unit IO-Link Display Module")
        item_number += 1
        #Tittle plus 2 Channels
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","HPU"],["equipment_count",""],["device_ID","DP1"],["device_description","Tunnel Evaporator Temperatures"]]
        addrow(item_number,globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],"")
        while i <= 2:
            locallist = [["io_start",i],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area_",""+"_"],["equipment_ID",""],["equipment_count",""],["device_ID",""],["device_description","IO-Link Converter Ch"+str(i)]]
            addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1]) 
            i += 1
            item_number += 1
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","MT"],["device_description","Hydraulic Power Unit Motor"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","ESR01"],["device_description","Hydraulic Power Unit Safety Contactor 1"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","ESR02"],["device_description","Hydraulic Power Unit Safety Contactor 2"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","MT01_RUN"],["device_description","Hydraulic Power Unit Motor Running"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1 
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","MT01_START"],["device_description","Hydraulic Power Unit Motor Start"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","TT01"],["device_description","Hydraulic Power Unit Motor Thermistor"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","LTT01"],["device_description","Hydraulic Power Unit Level Temperature"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","PT02"],["device_description","Hydraulic Power Unit Pressure"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","PT06"],["device_description","Hydraulic Power Pressure Filter Blocked"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","PT07"],["device_description","Hydraulic Power Return Filter Blocked"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","CV01"],["device_description","Hydraulic Power Unit Pressure Control"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","LS01"],["device_description","Hydraulic Power Unit Low level"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CTH"+"_"],["equipment_ID","PP"],["equipment_count",""],["device_ID","SV02"],["device_description","Hydraulic Power Unit Heating Bypass"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1


        addtitle("TUNNEL SAFETY")
        item_number += 1
        #TUNNEL SAFETY
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID",""],["equipment_count",""],["device_ID",""],["device_description","TUNNEL SAFETY DEVICES"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID",""],["equipment_count",""],["device_ID",""],["device_description","Entry and Exit Doors"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","OP01"],["device_description","Tunnel Entry Door 1"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SHS01_Ch1"],["device_description","Tunnel Entry Door 1 Removable Tag"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SHS01_Ch2"],["device_description","Tunnel Entry Door 1 Removable Tag"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SHS02_Ch1"],["device_description","Tunnel Entry Door 1 Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SHS02_Ch2"],["device_description","Tunnel Entry Door 1 Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","HS03"],["device_description","Tunnel Entry Door 1 Reset"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","IL03"],["device_description","Tunnel Entry Door 1 Reset"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","HS04"],["device_description","Tunnel Entry Door 1 Start"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","IL04"],["device_description","Tunnel Entry Door 1 Start"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","LK01"],["device_description","Tunnel Entry Door 1 Mag Lock 1"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SZS01_Ch1"],["device_description","Tunnel Entry Door 1 Mag Lock 1 Opened"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","SZS01_Ch2"],["device_description","Tunnel Entry Door 1 Mag Lock 1 Opened"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","01"],["device_ID","HS05"],["device_description","Tunnel Entry Door 1 Exit"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1

        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","OP01"],["device_description","Tunnel Entry Door 2"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SHS01_Ch1"],["device_description","Tunnel Entry Door 2 Removable Tag"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SHS01_Ch2"],["device_description","Tunnel Entry Door 2 Removable Tag"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SHS02_Ch1"],["device_description","Tunnel Entry Door 2 Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SHS02_Ch2"],["device_description","Tunnel Entry Door 2 Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","HS03"],["device_description","Tunnel Entry Door 2 Reset"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","IL03"],["device_description","Tunnel Entry Door 2 Reset"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","HS04"],["device_description","Tunnel Entry Door 2 Start"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","IL04"],["device_description","Tunnel Entry Door 2 Start"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","LK01"],["device_description","Tunnel Entry Door 2 Mag Lock 2"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SZS01_Ch1"],["device_description","Tunnel Entry Door 2 Mag Lock 2 Opened"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","SZS01_Ch2"],["device_description","Tunnel Entry Door 2 Mag Lock 2 Opened"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","02"],["device_ID","HS05"],["device_description","Tunnel Entry Door 2 Exit"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1

        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","03"],["device_ID","SZS01"],["device_description","Tunnel Exit Door 3"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","03"],["device_ID","SZS01_Ch1"],["device_description","Tunnel Exit Door 3"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","DR"],["equipment_count","03"],["device_ID","SZS01_Ch2"],["device_description","Tunnel Exit Door 3"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","OP"],["equipment_count","03"],["device_ID",""],["device_description","Tunnel Transfer End Lower Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","OP"],["equipment_count","SHS01_Ch1"],["device_ID",""],["device_description","Tunnel Transfer End Lower Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","OP"],["equipment_count","SHS01_Ch2"],["device_ID",""],["device_description","Tunnel Transfer End Lower Emergency Stop"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1

        addtitle("Light Stacks")
        item_number += 1

        #Light Stacks
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","OP"],["equipment_count","01"],["device_ID"," "],["device_description","Tunnel Enclosure Warning"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CT"+"_"],["equipment_ID","OP"],["equipment_count","02"],["device_ID"," "],["device_description","Tunnel Enclosure Warning"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1
        locallist = [["io_start",1],["itemcount",item_number],["engineer_name","Grayson Stillwell"],["area","CR"+"_"],["equipment_ID","OP"],["equipment_count","03"],["device_ID"," "],["device_description","Tunnel Control Room Warning"]]
        addrow(locallist[1][1],globallist[2][1],locallist[3][1],locallist[4][1],locallist[5][1],locallist[6][1],locallist[7][1])
        item_number += 1

        #Tunnel Load End
        addtitle("Tunnel Load End")
        item_number += 1
        
                
            
        
        
        
        wb.save("V.01/test_data.xlsx")

        return render_template('success.html', input1=input1, input2=input2, input3=input3,input4=input4, input5=input5, input6=input6,input7=input7, input8=input8, input9=input9,input10=input10, input11=input11, input12=input12, )
        
    return render_template('index.html')

@app.route('/download')
def download_file():
    return send_file("test_data.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="test_data.xlsx")

if __name__ == '__main__':
    app.run(debug=True)
